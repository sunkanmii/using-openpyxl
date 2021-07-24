import openpyxl

firstFile = r"company-in-nigeria.xlsx"
secondFile = r"company-in-rwanda.xlsx"

wbObj1 = openpyxl.load_workbook(firstFile)    
wbObj2 = openpyxl.load_workbook(secondFile)

wbk1 = wbObj1.worksheets[0]
wbk2 = wbObj2.worksheets[0]

names = []
for myRow in range(2, wbk2.max_row + 1):
    names.append(wbk2.cell(row=myRow, column=1).value.lower())
wbObj2.close

for rowNum in range(2, wbk1.max_row + 1):
    if wbk1.cell(row=rowNum, column=1).value.lower() in names:
	    wbk1.cell(row=rowNum, column=2).value = "Present"
    else:
	    wbk1.cell(row=rowNum, column=2).value = "Not present"
wbObj1.save(firstFile)
wbObj1.close